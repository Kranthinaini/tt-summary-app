import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

st.set_page_config(page_title="TT Summary Report", layout="wide")

st.title("TT Summary Report")

# Upload Files
credit_file = st.file_uploader(
    "Upload Credit Report",
    type=["xlsx"],
    key="credit"
)

db_file = st.file_uploader(
    "Upload DB Report",
    type=["xlsx"],
    key="db"
)

if credit_file is not None and db_file is not None:

    # Read Credit Report
    credit_df = pd.read_excel(credit_file)
    credit_df.columns = credit_df.columns.str.strip()

    # Read DB workbook
    xls = pd.ExcelFile(db_file)

    st.subheader("Select DB Sheets")

    vfs_sheet = st.selectbox(
        "Select VFS Sheet",
        xls.sheet_names
    )

    vivifi_sheet = st.selectbox(
        "Select VIVIFI Sheet",
        xls.sheet_names
    )

    # Read selected sheets
    vfs_df = pd.read_excel(db_file, sheet_name=vfs_sheet)
    vivifi_df = pd.read_excel(db_file, sheet_name=vivifi_sheet)

    # Combine
    db_df = pd.concat(
        [vfs_df, vivifi_df],
        ignore_index=True
    )

    db_df.columns = db_df.columns.str.strip()

    #st.write("DB Columns")
    #st.write(db_df.columns.tolist())

    #st.write("Credit Columns")
    #st.write(credit_df.columns.tolist())
    # Standardize Cluster Column

    credit_df.columns = credit_df.columns.str.strip()
    db_df.columns = db_df.columns.str.strip()

    credit_df["Cluster"] = credit_df["Cluster Name"]   


    start_date = st.date_input("From Date")
    end_date = st.date_input("To Date")
    start_ts = pd.to_datetime(start_date)
    end_ts = pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)

    today = pd.Timestamp.today().normalize()

    if st.button("Generate Report"):

        # -------------------------
        # Credit Dates
        # -------------------------
        credit_df["Submitted to TCM Date Time"] = pd.to_datetime(
            credit_df["Submitted to TCM Date Time"],
            errors="coerce"
        )

        credit_df["Latest Status Date"] = pd.to_datetime(
            credit_df["Latest Status Date"],
            errors="coerce"
        )

        credit_df["CCT Approved Date Time"] = pd.to_datetime(
            credit_df["CCT Approved Date Time"],
            errors="coerce"
        )

        # -------------------------
        # DB Date
        # -------------------------
        db_df["Disbursement Date"] = pd.to_datetime(
            db_df["Disbursement Date"],
            errors="coerce"
        )
        daily_tcm = (
            credit_df[
                credit_df["Submitted to TCM Date Time"].dt.normalize() == today
            ]
            .groupby("Cluster")
            .size()
            .rename("Daily_TCM_Submitted")
        )
        # Define reject statuses here
        reject_statuses = [
            "RCM Verified and Rejected",
            "RCM Rejected",
            "TCM Rejected",
            "CCT Rejected"
        ]

        today = pd.Timestamp.today().normalize()
        daily_reject = (
            credit_df[
                (credit_df["Latest Status Date"].dt.normalize() == today)
                &
                (credit_df["Final Status"].isin(reject_statuses))
            ]
            .groupby("Cluster")
            .size()
            .rename("Daily_Rejected")
        )

        daily_approved = (
            credit_df[
                credit_df["CCT Approved Date Time"].dt.normalize() == today
            ]
            .groupby("Cluster")
            .size()
            .rename("Daily_Approved")
        )

        daily_db = db_df[
            db_df["Disbursement Date"].dt.normalize() == today
        ]

        daily_db_count = (
            daily_db.groupby("Cluster")
            .size()
            .rename("Daily_DB_Count")
        )

        daily_db_amt = (
            daily_db.groupby("Cluster")["Disbursement Amount"]
            .sum()
            .rename("Daily_DB_Amount")
        )
        # -------------------------
        # TCM Submitted
        # -------------------------
        monthly_tcm = (
            credit_df[
                (credit_df["Submitted to TCM Date Time"] >= start_ts) &
                (credit_df["Submitted to TCM Date Time"] <= end_ts)
            ]
            .groupby("Cluster")
            .size()
            .rename("Monthly_TCM_Submitted")
        )

        monthly_reject = (
            credit_df[
                (credit_df["Latest Status Date"] >= start_ts) &
                (credit_df["Latest Status Date"] <= end_ts) &
                (credit_df["Final Status"].isin(reject_statuses))
            ]
            .groupby("Cluster")
            .size()
            .rename("Monthly_Rejected")
        )

        monthly_approved = (
            credit_df[
                (credit_df["CCT Approved Date Time"] >= start_ts) &
                (credit_df["CCT Approved Date Time"] <= end_ts)
            ]
            .groupby("Cluster")
            .size()
            .rename("Monthly_Approved")
        )
        db_df["Disbursement Date"] = pd.to_datetime(
            db_df["Disbursement Date"],
            errors="coerce"
        )

        db_filtered = db_df[
            (db_df["Disbursement Date"] >= start_ts) &
            (db_df["Disbursement Date"] <= end_ts)
        ]
        monthly_db_count = (
            db_filtered.groupby("Cluster")
            .size()
            .rename("Monthly_DB_Count")
        )

        monthly_db_amt = (
            db_filtered.groupby("Cluster")["Disbursement Amount"]
            .sum()
            .rename("Monthly_DB_Amount")
        )
        # -------------------------
        # Final Report
        # -------------------------
        report = pd.concat(
            [
                daily_tcm,
                daily_reject,
                daily_approved,
                daily_db_count,
                daily_db_amt,
                monthly_tcm,
                monthly_reject,
                monthly_approved,
                monthly_db_count,
                monthly_db_amt
            ],
            axis=1
        ).fillna(0).reset_index()

        num_cols = [
            "Daily_TCM_Submitted",
            "Daily_Rejected",
            "Daily_Approved",
            "Daily_DB_Count",
            "Daily_DB_Amount",
            "Monthly_TCM_Submitted",
            "Monthly_Rejected",
            "Monthly_Approved",
            "Monthly_DB_Count",
            "Monthly_DB_Amount"
        ]

        for col in num_cols:
            if col in report.columns:
                report[col] = report[col].fillna(0).astype(int)
        #st.subheader("TT Summary")
        #st.dataframe(result, use_container_width=True)
        total_row = pd.DataFrame({
            "Cluster": ["TOTAL"],

            "Daily_TCM_Submitted": [report["Daily_TCM_Submitted"].sum()],
            "Daily_Rejected": [report["Daily_Rejected"].sum()],
            "Daily_Approved": [report["Daily_Approved"].sum()],
            "Daily_DB_Count": [report["Daily_DB_Count"].sum()],
            "Daily_DB_Amount": [report["Daily_DB_Amount"].sum()],

            "Monthly_TCM_Submitted": [report["Monthly_TCM_Submitted"].sum()],
            "Monthly_Rejected": [report["Monthly_Rejected"].sum()],
            "Monthly_Approved": [report["Monthly_Approved"].sum()],
            "Monthly_DB_Count": [report["Monthly_DB_Count"].sum()],
            "Monthly_DB_Amount": [report["Monthly_DB_Amount"].sum()]
        })
        # Get Cluster-Territory mapping
        cluster_tt_credit = credit_df[
            ["Territory", "Cluster"]
        ].drop_duplicates()

        cluster_tt_db = db_df[
            ["Territory", "Cluster"]
        ].drop_duplicates()

        cluster_tt_map = pd.concat(
            [cluster_tt_credit, cluster_tt_db],
            ignore_index=True
        ).drop_duplicates(subset=["Cluster"])

        result = report.merge(
            cluster_tt_map,
            on="Cluster",
            how="left"
        )
        result["Territory"] = result["Territory"].fillna("UNKNOWN")

        #st.write("Columns After Merge")
        #st.write(result.columns.tolist())
            
        result = result.sort_values(
            ["Territory", "Cluster"]
        )
        final_rows = []

        for tt, grp in result.groupby("Territory", sort=False):

            final_rows.append(grp)

            total_row_tt = pd.DataFrame({
                "Territory": [tt],
                "Cluster": [f"TOTAL-{tt}"],

                "Daily_TCM_Submitted": [grp["Daily_TCM_Submitted"].sum()],
                "Daily_Rejected": [grp["Daily_Rejected"].sum()],
                "Daily_Approved": [grp["Daily_Approved"].sum()],
                "Daily_DB_Count": [grp["Daily_DB_Count"].sum()],
                "Daily_DB_Amount": [grp["Daily_DB_Amount"].sum()],

                "Monthly_TCM_Submitted": [grp["Monthly_TCM_Submitted"].sum()],
                "Monthly_Rejected": [grp["Monthly_Rejected"].sum()],
                "Monthly_Approved": [grp["Monthly_Approved"].sum()],
                "Monthly_DB_Count": [grp["Monthly_DB_Count"].sum()],
                "Monthly_DB_Amount": [grp["Monthly_DB_Amount"].sum()]
            })

            final_rows.append(total_row_tt)

            if len(final_rows) == 0:
                st.warning("No Territory data found. Check input files or mapping.")
                st.stop()

            if len(final_rows) == 0:
                st.error("No Territory data found. Check input files.")
                st.stop()

            final_report = pd.concat(final_rows, ignore_index=True)
        # Grand Total Row
        grand_total_row = pd.DataFrame({
            "Cluster": ["GRAND TOTAL"],

            "Daily_TCM_Submitted": [result["Daily_TCM_Submitted"].sum()],
            "Daily_Rejected": [result["Daily_Rejected"].sum()],
            "Daily_Approved": [result["Daily_Approved"].sum()],
            "Daily_DB_Count": [result["Daily_DB_Count"].sum()],
            "Daily_DB_Amount": [result["Daily_DB_Amount"].sum()],

            "Monthly_TCM_Submitted": [result["Monthly_TCM_Submitted"].sum()],
            "Monthly_Rejected": [result["Monthly_Rejected"].sum()],
            "Monthly_Approved": [result["Monthly_Approved"].sum()],
            "Monthly_DB_Count": [result["Monthly_DB_Count"].sum()],
            "Monthly_DB_Amount": [result["Monthly_DB_Amount"].sum()]
        })

        final_report = pd.concat(
            [final_report, grand_total_row],
            ignore_index=True
        )


        final_report = final_report[
            [
                "Cluster",

                "Daily_TCM_Submitted",
                "Daily_Rejected",
                "Daily_Approved",
                "Daily_DB_Count",
                "Daily_DB_Amount",

                "Monthly_TCM_Submitted",
                "Monthly_Rejected",
                "Monthly_Approved",
                "Monthly_DB_Count",
                "Monthly_DB_Amount"
            ]
        ]
        display_report = final_report.copy()

        display_report.columns = [
            "Cluster",

            "TCM Submitted",
            "Rejected",
            "Approved",
            "DB Count",
            "DB Amount",

            "TCM Submitted",
            "Rejected",
            "Approved",
            "DB Count",
            "DB Amount"
        ]

        #st.subheader("Cluster Summary with TT Totals")
        #st.dataframe(final_report, use_container_width=True)

        #st.subheader("Grand Total")
        #st.dataframe(total_row, use_container_width=True)

        output = BytesIO()

        # Dynamic Report Title
        report_title = (
            f"VFS Summary Report "
            f"({start_ts.strftime('%d-%m-%Y')} to {end_ts.strftime('%d-%m-%Y')})"
        )

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            export_report = final_report.copy()

            export_report.columns = [
                "Cluster",
                "TCM Submitted", "Rejected", "Approved", "DB Count", "DB Amount",
                "TCM Submitted", "Rejected", "Approved", "DB Count", "DB Amount"
            ]

            # Start data from row 3 because row 1 = title, row 2-3 = headers
            export_report.to_excel(
                writer,
                sheet_name="TT Summary",
                index=False,
                startrow=2
            )

            ws = writer.sheets["TT Summary"]

            # ==================================
            # IMPORT STYLES
            # ==================================
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

            # ==================================
            # COLORS
            # ==================================
            header_font = Font(color="FFFFFF", bold=True)

            cluster_fill = PatternFill("solid", fgColor="2F5597")
            daily_fill = PatternFill("solid", fgColor="1F4E78")
            monthly_fill = PatternFill("solid", fgColor="C65911")

            light_blue = PatternFill("solid", fgColor="D9E1F2")
            light_orange = PatternFill("solid", fgColor="FCE4D6")

            tt_fill = PatternFill("solid", fgColor="E0FFC2")
            grand_fill = PatternFill("solid", fgColor="548235")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # ==================================
            # REPORT TITLE
            # ==================================
# ==================================
# REPORT TITLE
# ==================================
            ws.merge_cells("A1:K1")

            ws["A1"] = report_title

            # Title Background Color
            ws["A1"].fill = PatternFill(
                "solid",
                fgColor="008080"      # Teal Color
            )

            # Title Font
            ws["A1"].font = Font(
                bold=True,
                size=16,
                color="FFFFFF"
            )

            # Title Alignment
            ws["A1"].alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            # Title Row Height
            ws.row_dimensions[1].height = 30

            # ==================================
            # MERGED HEADERS
            # ==================================
            ws.merge_cells("A2:A3")
            ws.merge_cells("B2:F2")
            ws.merge_cells("G2:K2")

            ws["A2"] = "CLUSTER"
            ws["B2"] = "DAILY REPORT"
            ws["G2"] = "MONTHLY REPORT"

            # Header Row Heights
            ws.row_dimensions[2].height = 22
            ws.row_dimensions[3].height = 22

            # ==================================
            # HEADER ROW 1 STYLE
            # ==================================
            ws["A2"].fill = cluster_fill
            ws["A2"].font = header_font
            ws["A2"].alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            for col in range(2, 7):
                c = ws.cell(row=2, column=col)
                c.fill = daily_fill
                c.font = header_font
                c.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            for col in range(7, 12):
                c = ws.cell(row=2, column=col)
                c.fill = monthly_fill
                c.font = header_font
                c.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # ==================================
            # HEADER ROW 2 STYLE
            # ==================================
            for col in range(1, 12):

                c = ws.cell(row=3, column=col)

                c.font = Font(bold=True)

                c.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

                if col == 1:
                    c.fill = cluster_fill

                elif 2 <= col <= 6:
                    c.fill = light_blue

                elif 7 <= col <= 11:
                    c.fill = light_orange

            # ==================================
            # DATA ROW STYLING
            # ==================================
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row):

                val = str(row[0].value)

                if val.startswith("TOTAL-"):
                    for cell in row:
                        cell.fill = tt_fill
                        cell.font = Font(bold=True)

                elif val == "GRAND TOTAL":
                    for cell in row:
                        cell.fill = grand_fill
                        cell.font = Font(
                            color="FFFFFF",
                            bold=True
                        )

            # ==================================
            # BORDERS
            # ==================================
            for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=1,
                max_col=11
            ):
                for cell in row:
                    cell.border = thin_border

            # ==================================
            # FREEZE PANES
            # ==================================
            ws.freeze_panes = "A4"

            # ==================================
            # AUTO COLUMN WIDTH
            # ==================================
            from openpyxl.utils import get_column_letter

            for col_idx in range(1, ws.max_column + 1):

                max_length = 0
                col_letter = get_column_letter(col_idx)

                for row in ws.iter_rows(
                    min_col=col_idx,
                    max_col=col_idx
                ):
                    for cell in row:
                        try:
                            if cell.value is not None:
                                max_length = max(
                                    max_length,
                                    len(str(cell.value))
                                )
                        except:
                            pass

                ws.column_dimensions[col_letter].width = max_length + 3

            # ==================================
            # CLUSTER COLUMN EXTRA WIDTH
            # ==================================
            ws.column_dimensions["A"].width = 35

            # ==================================
            # CENTER ALIGN ALL DATA
            # ==================================
            for row in ws.iter_rows(
                min_row=4,
                max_row=ws.max_row,
                min_col=1,
                max_col=11
            ):
                for cell in row:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )
        output.seek(0)

        st.success("TT Summary report generated successfully.")
        st.download_button(
            label="Download TT Summary",
            data=output.getvalue(),
            file_name="TT_Summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        #python -m streamlit run newreport.py
        #python -m streamlit run newreport.py